import streamlit as st
import pandas as pd
from pyspark.sql import SparkSession
from pyspark.sql.functions import col, lit, concat, monotonically_increasing_id
import tempfile
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
from datetime import datetime
from typing import Dict, List, Tuple, Any
from anthropic import Anthropic
import io
import numpy as np
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO

# Set up logging and Anthropic client
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
anthropic = Anthropic()

# Initialize Spark Session
@st.cache_resource
def create_spark_session():
    return SparkSession.builder \
        .appName("FileTransformer") \
        .config("spark.driver.memory", "4g") \
        .getOrCreate()

def read_file(uploaded_file: Any) -> pd.ExcelFile:
    """Read different file formats and return as ExcelFile object"""
    file_type = uploaded_file.name.split('.')[-1].lower()
    
    if file_type in ['xlsx', 'xls', 'xlsm', 'xlsb', 'odf', 'ods', 'odt']:
        return pd.ExcelFile(uploaded_file)
    elif file_type == 'csv':
        buffer = io.BytesIO()
        df = pd.read_csv(uploaded_file)
        with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            writer.close()
        buffer.seek(0)
        return pd.ExcelFile(buffer)
    else:
        raise ValueError(f"Unsupported file format: {file_type}")

def process_file_by_sections(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """Process file section by section, maintaining the sequence"""
    section_sequence = [
        "BEASLEY DAY SHIFT PRODUCTION",
        "SOS SECTION",
        "Amazon ( Holweg + Garant ) Machines",
        "GARANT SECTION",
        "HANDLE SECTION",
        "MATADOR SECTION",
        "WICKETTING SECTION",
        "SHEETER SECTION",
        "DAILY PRODUCTION REPORT (NIGHT SHIFT)",
        "BEASLEY SECTION",
        "MATADOR SECTION",
        "SOS SECTION",
        "GARANT SECTION",
        "SHEETER SECTION",
        "Amazon ( Holweg + Garant ) Machines",
        "HANDLE SECTION"
    ]
    
    # Define section identifiers with variations
    section_identifiers = {
        "BEASLEY DAY SHIFT PRODUCTION": ["BEASLEY", "B-"],
        "SOS SECTION": ["SOS", "CH-"],
        "Amazon ( Holweg + Garant ) Machines": ["AMAZON", "HOLWEG", "GARANT MACHINES", "AMAZON ( HOLWEG", "HOLWEG + GARANT"],
        "GARANT SECTION": ["GARANT SECTION"],
        "HANDLE SECTION": ["HANDLE"],
        "MATADOR SECTION": ["MATADOR"],
        "WICKETTING SECTION": ["WICKETTING"],
        "SHEETER SECTION": ["SHEETER"],
        "BEASLEY SECTION": ["BEASLEY", "B-"],
        "DAILY PRODUCTION REPORT (NIGHT SHIFT)": ["NIGHT SHIFT", "NIGHT PRODUCTION"]
    }
    
    processed_sections = []
    current_section = None
    section_data = []
    is_night_shift = False
    
    def match_section(text: str, section_name: str) -> bool:
        """Helper function to match section text with identifiers"""
        if section_name not in section_identifiers:
            return False
        return any(identifier in text for identifier in section_identifiers[section_name])
    
    def get_section_for_machine(machine_code: str) -> Tuple[str, bool]:
        """Determine section and shift based on machine code"""
        machine_code = machine_code.upper()
        is_night = machine_code.endswith("-N")
        
        if machine_code.startswith("B-"):
            return ("BEASLEY SECTION" if is_night else "BEASLEY DAY SHIFT PRODUCTION", is_night)
        elif machine_code.startswith("CH-"):
            return ("SOS SECTION", is_night)
        return (None, is_night)
    
    for idx, row in df.iterrows():
        row_text = ' '.join(str(cell).strip() for cell in row if pd.notna(cell)).upper()
        
        if not row_text.strip():
            continue
        
        # Check for night shift marker
        if "NIGHT SHIFT" in row_text or "NIGHT PRODUCTION" in row_text:
            is_night_shift = True
            if current_section and section_data:
                processed_sections.append({
                    'name': current_section,
                    'data': pd.DataFrame(section_data),
                    'shift': 'Day'
                })
                section_data = []
            current_section = None
            continue
        
        # Check for section headers
        for section in section_sequence:
            if match_section(row_text, section):
                if current_section and section_data:
                    processed_sections.append({
                        'name': current_section,
                        'data': pd.DataFrame(section_data),
                        'shift': 'Night' if is_night_shift else 'Day'
                    })
                    section_data = []
                current_section = section
                break
        
        # Special handling for Amazon section
        if "AMAZON" in row_text and "HOLWEG" in row_text:
            if current_section and section_data:
                processed_sections.append({
                    'name': current_section,
                    'data': pd.DataFrame(section_data),
                    'shift': 'Night' if is_night_shift else 'Day'
                })
                section_data = []
            current_section = "Amazon ( Holweg + Garant ) Machines"
            continue
        
        # Process data rows
        if not any(header in row_text for header in ['MACHINE NO', 'SUPERVISOR', 'SUMMARY', 'TOTAL']):
            # Check for machine codes to determine section
            first_cell = str(row.iloc[0]).upper().strip()
            section_from_machine, machine_night_shift = get_section_for_machine(first_cell)
            
            if section_from_machine:
                # If we find a new section from machine code, save current section data
                if current_section and current_section != section_from_machine and section_data:
                    processed_sections.append({
                        'name': current_section,
                        'data': pd.DataFrame(section_data),
                        'shift': 'Night' if is_night_shift else 'Day'
                    })
                    section_data = []
                
                current_section = section_from_machine
                is_night_shift = machine_night_shift or is_night_shift
                section_data.append(row)
            elif current_section:
                section_data.append(row)
    
    # Add last section
    if current_section and section_data:
        processed_sections.append({
            'name': current_section,
            'data': pd.DataFrame(section_data),
            'shift': 'Night' if is_night_shift else 'Day'
        })
    
    return processed_sections

def process_section_data(df_section: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
    """Process section DataFrame by cleaning and standardizing data"""
    df_section = df_section.copy()
    
    # Filter out summary rows, headers, and "Operator Cost per tonn"
    df_section = df_section[
        ~(
            (df_section.iloc[:, 0].astype(str).str.contains('Machine No|Summary|Total|Operator Cost per tonn', case=False, na=False)) |
            (df_section.iloc[:, 1].astype(str).str.contains('Supervisor|Summary|Total|Operator Cost per tonn', case=False, na=False)) |
            (df_section.iloc[:, 2].astype(str).str.contains('NAME|Summary|Total|Operator Cost per tonn', case=False, na=False)) |
            # Check all columns for "Operator Cost per tonn"
            df_section.apply(lambda x: x.astype(str).str.contains('Operator Cost per tonn', case=False, na=False)).any(axis=1)
        )
    ]
    
    # Remove completely empty rows
    df_section = df_section.dropna(how='all')
    
    # Rename columns
    for i, col in enumerate(columns):
        if i < len(df_section.columns):
            df_section.rename(columns={df_section.columns[i]: col}, inplace=True)
    
    # Clean numeric columns
    clean_numeric_columns(df_section)
    
    # Remove rows where Machine_No is missing or invalid
    df_section = df_section[
        df_section['Machine_No'].notna() & 
        (df_section['Machine_No'].astype(str).str.strip() != '')
    ]
    
    return df_section

def clean_numeric_columns(df: pd.DataFrame) -> None:
    """Clean numeric columns in the DataFrame"""
    numeric_columns = [
        'Hours', 'Operator_Cost', 'Net_Weight', 'Per_Pack', 'Bag_Produce',
        'Packet_Produce', 'In_Kgs', 'target_Bag_Produce', 'Pkt', 'KG_Target',
        'Pkt_Var_%', 'KG_Variance_%'
    ]
    
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

def add_summary_row(df: pd.DataFrame) -> pd.DataFrame:
    """Add a summary row to the dataframe with specified calculations"""
    summary_row = {
        'Machine_No': f"Total Unique Machines: {df['Machine_No'].nunique()}",
        'Supervisor': 'SUMMARY',
        'Hours': df['Hours'].sum(),
        'Operator_Cost': df['Operator_Cost'].sum(),
        'NAME': 'TOTAL',
        'SAP_Code': '',
        'Net_Weight': df['Net_Weight'].mean(),  # Average net weight
        'Size': '',
        'Material_Description': '',
        'Per_Pack': df['Per_Pack'].sum(),
        'Bag_Produce': df['Bag_Produce'].sum(),
        'Packet_Produce': df['Packet_Produce'].sum(),
        'In_Kgs': df['In_Kgs'].sum(),
        'target_Bag_Produce': df['target_Bag_Produce'].sum(),
        'Pkt': df['Pkt'].sum(),
        'KG_Target': df['KG_Target'].sum()
    }
    
    # Create summary DataFrame
    summary_df = pd.DataFrame([summary_row])
    
    # Add any additional columns that might be in the original DataFrame
    for col in df.columns:
        if col not in summary_df.columns:
            summary_df[col] = ''
    
    # Combine original data with summary row
    return pd.concat([df, summary_df], ignore_index=True)

def transform_to_capacity_report(sections_with_summary: Dict[Tuple[str, str], pd.DataFrame]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Transform section analysis data into capacity report format matching EFESO template"""
    # Use section sequence and identifiers from template.py
    section_sequence = [
        "BEASLEY DAY SHIFT PRODUCTION",
        "SOS SECTION",
        "Amazon ( Holweg + Garant ) Machines",
        "GARANT SECTION",
        "HANDLE SECTION",
        "MATADOR SECTION",
        "WICKETTING SECTION",
        "SHEETER SECTION",
        "DAILY PRODUCTION REPORT (NIGHT SHIFT)",
        "BEASLEY SECTION",
        "MATADOR SECTION",
        "SOS SECTION",
        "GARANT SECTION",
        "SHEETER SECTION",
        "Amazon ( Holweg + Garant ) Machines",
        "HANDLE SECTION"
    ]
    
    # Define section identifiers with variations
    section_identifiers = {
        "BEASLEY DAY SHIFT PRODUCTION": ["BEASLEY", "B-"],
        "SOS SECTION": ["SOS", "CH-"],
        "Amazon ( Holweg + Garant ) Machines": ["AMAZON", "HOLWEG", "GARANT MACHINES", "AMAZON ( HOLWEG", "HOLWEG + GARANT"],
        "GARANT SECTION": ["GARANT SECTION"],
        "HANDLE SECTION": ["HANDLE"],
        "MATADOR SECTION": ["MATADOR"],
        "WICKETTING SECTION": ["WICKETTING"],
        "SHEETER SECTION": ["SHEETER"],
        "BEASLEY SECTION": ["BEASLEY", "B-"],
        "DAILY PRODUCTION REPORT (NIGHT SHIFT)": ["NIGHT SHIFT", "NIGHT PRODUCTION"]
    }
    
    # Get unique sections for the report
    unique_sections = []
    seen_sections = set()
    for section in section_sequence:
        base_section = section.split("(")[0].strip()  # Remove shift indicators
        if base_section not in seen_sections and base_section != "DAILY PRODUCTION REPORT":
            unique_sections.append(base_section)
            seen_sections.add(base_section)
    
    # Add Total to the sections
    processes = unique_sections + ["Total"]
    
    # Initialize data structures for each section
    combined_sections = {process: {
        'machines': 0,
        'weekly_volume': 0,
        'actual_hours': 0,
        'efficiency_sum': 0,
        'efficiency_count': 0,
        'total_manhours': 0,
        'total_target': 0
    } for process in processes[:-1]}  # Exclude Total
    
    # Helper function to get base section name
    def get_base_section(section_name: str) -> str:
        for base_section, identifiers in section_identifiers.items():
            if any(identifier in section_name.upper() for identifier in identifiers):
                return base_section.split("(")[0].strip()
        return section_name.split("(")[0].strip()
    
    # Process data from sections
    for (section_name, shift), data in sections_with_summary.items():
        base_section = get_base_section(section_name)
        if base_section in combined_sections:
            section = combined_sections[base_section]
            section['machines'] = max(section['machines'], data['Machine_No'].nunique())
            section['weekly_volume'] += data['Packet_Produce'].sum()
            section['actual_hours'] += data['Hours'].sum()
            section['total_target'] += data['Pkt'].sum()
            efficiency = (data['Packet_Produce'].sum() / data['Pkt'].sum() * 100) if data['Pkt'].sum() > 0 else 0
            section['efficiency_sum'] += efficiency
            section['efficiency_count'] += 1
            section['total_manhours'] += data['Hours'].sum()
    
    # Create capacity and labour data structures
    capacity_data = {
        'Process': processes,
        '# Mach. Avail.': [],
        'Production Volume (Weekly)': [],
        'Meas. Unit': [],
        'Value Adding Mc Hours/Week/Mc': [],
        'Ref Speed (Meas. Unit per min)': [],
        'Actual OEE': [],
        'Actual Mc Hours/Week/Mc': [],
        'Saturation vs. 110 hrs/wk': []
    }
    
    labour_data = {
        'Dir op/mach/shift': [],
        'Required Manhours/Week': [],
        'Actual Manhours/Week': [],
        'Org. Losses': [],
        '% Overtime': [],
        '% Absence': [],
        'Actual No of Dir Ops': []
    }
    
    # Calculate totals
    total_machines = 0
    total_volume = 0
    total_manhours_required = 0
    total_manhours_actual = 0
    total_ops = 0
    
    # Fill data for each process
    for process in processes[:-1]:
        data = combined_sections[process]
        machines = data['machines']
        total_machines += machines
        
        weekly_volume = data['weekly_volume']
        total_volume += weekly_volume
        
        actual_hours = data['actual_hours']
        target_volume = data['total_target']
        
        # Capacity calculations
        value_adding_hours = actual_hours / machines if machines > 0 else 0
        ref_speed = weekly_volume / (actual_hours * 60) if actual_hours > 0 else 0
        oee = (weekly_volume / target_volume * 100) if target_volume > 0 else 0
        
        # Labour calculations
        dir_per_shift = 2.0 if process == "SHEETER SECTION" else 1.0
        required_manhours = machines * 40  # 40 hours per machine per week
        actual_manhours = data['total_manhours']
        org_losses = ((actual_manhours - required_manhours) / required_manhours) if required_manhours > 0 else 0
        actual_ops = machines * dir_per_shift
        
        # Update totals
        total_manhours_required += required_manhours
        total_manhours_actual += actual_manhours
        total_ops += actual_ops
        
        # Add to capacity data
        capacity_data['# Mach. Avail.'].append(machines)
        capacity_data['Production Volume (Weekly)'].append(weekly_volume)
        capacity_data['Meas. Unit'].append('kg' if process == "GARANT SECTION" else 'pk')
        capacity_data['Value Adding Mc Hours/Week/Mc'].append(value_adding_hours)
        capacity_data['Ref Speed (Meas. Unit per min)'].append(ref_speed)
        capacity_data['Actual OEE'].append(oee)
        capacity_data['Actual Mc Hours/Week/Mc'].append(60)
        capacity_data['Saturation vs. 110 hrs/wk'].append(0.55)
        
        # Add to labour data
        labour_data['Dir op/mach/shift'].append(dir_per_shift)
        labour_data['Required Manhours/Week'].append(required_manhours)
        labour_data['Actual Manhours/Week'].append(actual_manhours)
        labour_data['Org. Losses'].append(org_losses)
        labour_data['% Overtime'].append(0.03)
        labour_data['% Absence'].append(0.02)
        labour_data['Actual No of Dir Ops'].append(actual_ops)
    
    # Add totals row
    capacity_data['# Mach. Avail.'].append(total_machines)
    capacity_data['Production Volume (Weekly)'].append(total_volume)
    capacity_data['Meas. Unit'].append('')
    capacity_data['Value Adding Mc Hours/Week/Mc'].append('')
    capacity_data['Ref Speed (Meas. Unit per min)'].append('')
    avg_oee = sum(capacity_data['Actual OEE'][:-1]) / len(processes[:-1]) if processes[:-1] else 0
    capacity_data['Actual OEE'].append(avg_oee)
    capacity_data['Actual Mc Hours/Week/Mc'].append('')
    capacity_data['Saturation vs. 110 hrs/wk'].append(0.55)
    
    labour_data['Dir op/mach/shift'].append('')
    labour_data['Required Manhours/Week'].append(total_manhours_required)
    labour_data['Actual Manhours/Week'].append(total_manhours_actual)
    labour_data['Org. Losses'].append((total_manhours_actual - total_manhours_required) / total_manhours_required if total_manhours_required > 0 else 0)
    labour_data['% Overtime'].append(0.03)
    labour_data['% Absence'].append(0.02)
    labour_data['Actual No of Dir Ops'].append(total_ops)
    
    return pd.DataFrame(capacity_data), pd.DataFrame(labour_data)

def export_capacity_report(capacity_df: pd.DataFrame, labour_df: pd.DataFrame, file_date: str = None) -> bytes:
    """Export capacity report in EFESO template format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Process Data"
    
    # Insert logo
    ws.insert_rows(1)
    ws.merge_cells('A1:B1')
    logo_cell = ws['A1']
    logo_cell.value = "EFESO Consulting"
    logo_cell.font = Font(bold=True)
    
    # Insert date header
    ws.insert_rows(2, 1)
    date_text = f"Start Point ({file_date})" if file_date else "Start Point"
    ws['A3'] = date_text
    ws.merge_cells('A3:P3')
    ws['A3'].fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
    ws['A3'].font = Font(bold=True)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Add section headers
    ws['A4'] = "Process"
    ws['B4'] = "CAPACITY"
    ws['J4'] = "LABOUR"
    
    # Merge cells for sections
    ws.merge_cells('B4:I4')
    ws.merge_cells('J4:P4')
    
    # Style section headers
    for cell in [ws['A4'], ws['B4'], ws['J4']]:
        cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Write column headers and data
    capacity_headers = ['Process'] + list(capacity_df.columns[1:])
    labour_headers = list(labour_df.columns)
    
    # Write headers
    for idx, header in enumerate(capacity_headers + labour_headers, 1):
        ws.cell(row=5, column=idx, value=header)
    
    # Write data
    for row_idx, (cap_row, lab_row) in enumerate(zip(capacity_df.values, labour_df.values), 6):
        for col_idx, value in enumerate(cap_row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        for col_idx, value in enumerate(lab_row, len(cap_row) + 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Style the sheet
    style_excel_sheet(wb)
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def style_excel_sheet(wb):
    """Style the Excel worksheet according to EFESO template"""
    ws = wb.active
    
    # Style borders
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for row in ws.rows:
        for cell in row:
            cell.border = thin_border
            if cell.row > 5:  # Don't center the headers
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Style process names
    light_blue = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
    for row in range(6, ws.max_row + 1):  # Start from row 6 (first data row)
        if ws[f'A{row}'].value in ['Beasley', 'SOS', 'Film Front', 'Sheeter', 'Handle', 'Printer', 'Wicketing']:
            ws[f'A{row}'].fill = light_blue
            ws[f'A{row}'].font = Font(color='0000FF')  # Blue text
    
    # Style totals row
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for cell in ws[ws.max_row]:
        if cell.value is not None:
            cell.fill = yellow_fill
    
    # Format percentage cells
    percentage_columns = ['% Overtime', '% Absence']
    for col in ws.columns:
        header = col[4].value  # Get column header
        if header in percentage_columns:
            for cell in col[5:]:  # Start after header
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0%'
    
    # Set column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

def main():
    st.set_page_config(
        page_title="Production Analysis Dashboard",
        page_icon="📊",
        layout="wide"
    )
    
    st.title("Production Analysis Dashboard")
    
    # Initialize sections_with_summary at the start
    sections_with_summary = {}
    
    # File upload and configuration section
    with st.sidebar:
        st.header("Upload Settings")
        batch_size = st.number_input(
            "Batch Size (sheets per process)",
            min_value=1,
            max_value=1000,
            value=50,
            help="Number of sheets to process in each batch"
        )
        
        show_progress = st.checkbox("Show Detailed Progress", value=True)
    
    uploaded_files = st.file_uploader(
        "Upload production files (Excel/CSV)",
        type=['xlsx', 'xls', 'csv'],
        accept_multiple_files=True
    )
    
    if uploaded_files:
        try:
            # Initialize progress tracking
            if show_progress:
                progress_bar = st.progress(0)
                status_text = st.empty()
            
            # Count total sheets
            total_sheets = 0
            file_info = []
            for file in uploaded_files:
                try:
                    excel_file = read_file(file)
                    n_sheets = len(excel_file.sheet_names)
                    total_sheets += n_sheets
                    file_info.append((file, n_sheets))
                except Exception as e:
                    logger.warning(f"Error counting sheets in {file.name}: {e}")
                finally:
                    file.seek(0)
            
            if show_progress:
                st.info(f"Found {total_sheets} sheets in {len(uploaded_files)} files")
            
            # Process files in batches
            all_sections_data = []
            processed_sheets = 0
            
            for file, n_sheets in file_info:
                try:
                    excel_file = read_file(file)
                    
                    # Process sheets in batches
                    for i in range(0, n_sheets, batch_size):
                        batch_sheets = excel_file.sheet_names[i:i + batch_size]
                        
                        for sheet_name in batch_sheets:
                            if show_progress:
                                status_text.text(f"Processing {file.name} - Sheet: {sheet_name}")
                                progress_bar.progress(processed_sheets / total_sheets)
                            
                            try:
                                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                                sections = process_file_by_sections(df)
                                
                                for section in sections:
                                    if not section['data'].empty:
                                        columns = [
                                            'Machine_No', 'Supervisor', 'Hours', 'Operator_Cost',
                                            'NAME', 'SAP_Code', 'Net_Weight', 'Size',
                                            'Material_Description', 'Per_Pack', 'Bag_Produce',
                                            'Packet_Produce', 'In_Kgs', 'target_Bag_Produce',
                                            'Pkt', 'KG_Target'
                                        ]
                                        processed_df = process_section_data(section['data'], columns)
                                        processed_df['Section'] = section['name']
                                        processed_df['Shift'] = section['shift']
                                        processed_df['File'] = file.name
                                        processed_df['Sheet'] = sheet_name
                                        all_sections_data.append(processed_df)
                                
                                processed_sheets += 1
                                
                            except Exception as e:
                                logger.warning(f"Error processing sheet {sheet_name} in {file.name}: {e}")
                                continue
                            
                except Exception as e:
                    logger.warning(f"Error processing file {file.name}: {e}")
                    continue
            
            if show_progress:
                progress_bar.progress(1.0)
                status_text.text("Processing complete!")
            
            if all_sections_data:
                # Create final dataframe
                final_df = pd.concat(all_sections_data, ignore_index=True)
                
                # Group by section and shift
                grouped = final_df.groupby(['Section', 'Shift'])
                
                # Create sections_with_summary dictionary
                sections_with_summary = {
                    (section_name, shift): group_data
                    for (section_name, shift), group_data in grouped
                }
                
                # Display summary statistics
                st.subheader("Processing Summary:")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Files Processed", len(uploaded_files))
                with col2:
                    st.metric("Sheets Processed", processed_sheets)
                with col3:
                    st.metric("Sections Found", len(grouped))
                
                # Display sections and their data with summary rows
                st.subheader("Sections and Their Data:")
                
                for (section_name, shift), group_data in grouped:
                    # Add summary row to the group data
                    group_with_summary = add_summary_row(group_data)
                    sections_with_summary[(section_name, shift)] = group_with_summary
                    
                    with st.expander(f"{section_name} ({shift})", expanded=True):
                        st.dataframe(
                            group_with_summary.drop(['Section', 'Shift'], axis=1),
                            use_container_width=True
                        )
                
                # Download options
                st.subheader("Download Options")
                col1, col2 = st.columns(2)
                
                # Prepare Excel file with multiple sheets including summary rows
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                    # Summary sheet
                    summary_data = {
                        'Section': [],
                        'Shift': [],
                        'Total Unique Machines': [],
                        'Total Hours': [],
                        'Total Operator Cost': [],
                        'Total Per Pack': [],
                        'Total Bag Produce': [],
                        'Total Packet Produce': [],
                        'Total In Kgs': [],
                        'Total Target Bag Produce': [],
                        'Total Pkt': [],
                        'Total KG Target': []
                    }
                    
                    # Write each section to a separate sheet
                    for (section_name, shift), group_data in sections_with_summary.items():
                        # Add to summary
                        summary_data['Section'].append(section_name)
                        summary_data['Shift'].append(shift)
                        summary_data['Total Unique Machines'].append(group_data['Machine_No'].nunique())
                        summary_data['Total Hours'].append(group_data['Hours'].sum())
                        summary_data['Total Operator Cost'].append(group_data['Operator_Cost'].sum())
                        summary_data['Total Per Pack'].append(group_data['Per_Pack'].sum())
                        summary_data['Total Bag Produce'].append(group_data['Bag_Produce'].sum())
                        summary_data['Total Packet Produce'].append(group_data['Packet_Produce'].sum())
                        summary_data['Total In Kgs'].append(group_data['In_Kgs'].sum())
                        summary_data['Total Target Bag Produce'].append(group_data['target_Bag_Produce'].sum())
                        summary_data['Total Pkt'].append(group_data['Pkt'].sum())
                        summary_data['Total KG Target'].append(group_data['KG_Target'].sum())
                        
                        # Write section data with summary row
                        sheet_name = f"{section_name[:20]}_{shift}"
                        group_data.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # Format sheet
                        worksheet = writer.sheets[sheet_name]
                        header_format = writer.book.add_format({
                            'bold': True,
                            'bg_color': '#D3D3D3'
                        })
                        summary_format = writer.book.add_format({
                            'bold': True,
                            'bg_color': '#FFE4B5'  # Light orange for summary row
                        })
                        
                        # Format headers
                        for col_num, value in enumerate(group_data.columns.values):
                            worksheet.write(0, col_num, value, header_format)
                        
                        # Format summary row
                        last_row = len(group_data)
                        for col_num in range(len(group_data.columns)):
                            worksheet.write(last_row, col_num, 
                                         group_data.iloc[-1][group_data.columns[col_num]], 
                                         summary_format)
                    
                    # Write summary sheet
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)
                    
                    # Format summary sheet
                    summary_sheet = writer.sheets['Summary']
                    for col_num, value in enumerate(summary_df.columns.values):
                        summary_sheet.write(0, col_num, value, header_format)
                
                with col1:
                    st.download_button(
                        label="Download Full Report (Excel)",
                        data=buffer.getvalue(),
                        file_name="production_data_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                with col2:
                    # Prepare CSV of all data including summary rows
                    csv_buffer = io.StringIO()
                    pd.concat(sections_with_summary.values(), ignore_index=True).to_csv(csv_buffer, index=False)
                    st.download_button(
                        label="Download All Data (CSV)",
                        data=csv_buffer.getvalue(),
                        file_name="production_data.csv",
                        mime="text/csv"
                    )
            else:
                st.warning("No data was processed. Please check your input files.")
                
        except Exception as e:
            st.error(f"Error processing files: {str(e)}")
            logger.error(f"Processing error: {str(e)}", exc_info=True)

    # Add capacity report generation section
    with st.expander("Generate Capacity Report", expanded=False):
        st.info("Generate a capacity report in EFESO template format")
        
        # Get the date from the first file name if available
        file_date = None
        if uploaded_files:
            try:
                first_file = uploaded_files[0].name
                # Try to extract date from filename (assuming format contains YYYYMMDD)
                import re
                date_match = re.search(r'(\d{8})', first_file)
                if date_match:
                    date_str = date_match.group(1)
                    file_date = datetime.strptime(date_str, '%Y%m%d').strftime('%Y-%m-%d')
            except:
                pass
        
        if sections_with_summary:
            try:
                # Generate capacity report
                capacity_df, labour_df = transform_to_capacity_report(sections_with_summary)
                
                # Display preview
                st.subheader("Capacity Report Preview")
                col1, col2 = st.columns(2)
                with col1:
                    st.write("Capacity Data")
                    st.dataframe(capacity_df)
                with col2:
                    st.write("Labour Data")
                    st.dataframe(labour_df)
                
                # Export button
                report_bytes = export_capacity_report(capacity_df, labour_df, file_date)
                st.download_button(
                    label="Download Capacity Report (Excel)",
                    data=report_bytes,
                    file_name=f"capacity_report_{file_date or 'generated'}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
            except Exception as e:
                st.error(f"Error generating capacity report: {str(e)}")
                logger.error(f"Capacity report generation error: {str(e)}", exc_info=True)
        else:
            st.warning("No data available for capacity report generation. Please upload and process files first.")

if __name__ == "__main__":
    main()